#!/usr/bin/env python
"""
Generate a justification sheet for an employee and a time period.

The process is composed of different steps:
    1. It starts with a project breakdown that lists, for each projet, the 
       employee projected dedication by year.
    2. This yearly projected dedication is then refined into a projected 
       dedication split into the given justification interval. For instance if
       the justification interval is set to 'month' the yearly projected 
       dedication will be split evenly into 12 intervals. Conversly if the 
       justification interval is set to 'week' the yearly projected 
       dedication will be split evenly into 52 intervals. Note that the 
       refined projected dedication can be adjusted by the user and picked 
       up by the script using the '--use-justification-hints' parameter.
       This parameter instruct the script to load the justified dedication 
       from a previsouly generated report and use them to create the new 
       report.
    3. Once the projected dedication has been refined for each justification
       interval, the script will start filling in each day of said interval 
       with the projects' projected dedication that are opened for that
       specific day, taking care of not breaching the limits imposed at 
       different level (yearly, monthly, by project, etc.), and also making
       sure that potential justification hints are respected for that project.
    4. Once the dedication has been established for the justification interval
       the script generate a report with the details of the justification for 
       the given justification interval, as well as a daily justification 
       sheet.

NOTE: The script relies heavily on the fact that the input Excel file has a 
      valid format. Do not modify the input Excel format without changing the
      corresponding parameter in the 'defauls.yaml' file.

NOTE: Use the 'resources/project-breakdown.xlsx' as a reference for the input 
      Excel file. 

SYNOPSIS:

# Generate a pristine justification sheet for John Smith for the current month
?> distribute_hours.py --employee "John Smith"

# Generate a pristine justification sheet for John Smith for June 2023
?> distribute_hours.py --employee "John Smith" --justification-period 2023-06-24

"""

try:
    from datetime import date, datetime, timedelta
    from dateutil.relativedelta import *
    from dotenv import load_dotenv
    from enum import Enum
    from functools import cmp_to_key

    import time
    import calendar
    import coloredlogs
    import locale
    import logging
    import math
    import os
    import re
    import sys
    import yaml
    from argparse import ArgumentParser, RawTextHelpFormatter
    from pprint import pprint, pformat
    from workalendar.europe import Aragon
    import openpyxl
    from openpyxl.formatting import Rule
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.styles import NamedStyle, Font, Border, Side, Color, PatternFill, Protection
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.cell.cell import Cell
    from openpyxl.packaging.custom import (BoolProperty, DateTimeProperty, FloatProperty, IntProperty, LinkProperty, StringProperty, CustomPropertyList)
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import quote_sheetname, absolute_coordinate, get_column_letter, column_index_from_string
except ModuleNotFoundError as e:
    print(f"{e}. Did you load your environment?")
    sys.exit(1)

load_dotenv()
logger = logging.getLogger()

DEFAULT_FMT = '%Y-%m-%d'
CELL_FMT = "0.0"  # "[h]:mm"
PREFERRED_FMT = 'decimal'
MONTH_INTERVAL = 'month'
WEEK_INTERVAL = 'week'
DEFAULT_LOCALE = 'es_ES'

# The row containing the interval id
INTERVAL_HEADER_ROW = 6
# The first row containing project information
PRJ_1ST_ROW = 7
PRJ_NAME_COL = 'A'
PRJ_NAME_COL_NUM = column_index_from_string(PRJ_NAME_COL)
# The column containing the projected dedication for the given project
PROJECTED_COL = 'B'
# The column containing the Excel SUM of justified dedication for the given project
JUSTIFIED_COL = 'C'
# The column containing the justified dedication for the given month / week
MONTH_1ST_COL = 'D'
MONTH_1ST_COL_NUM = column_index_from_string(MONTH_1ST_COL)
MONTH_LST_COL = get_column_letter(MONTH_1ST_COL_NUM + 12 - 1)
MONTH_LST_COL_NUM = column_index_from_string(MONTH_LST_COL)

WEEK_1ST_COL = 'D'
WEEK_1ST_COL_NUM = column_index_from_string(WEEK_1ST_COL)
WEEK_LST_COL = get_column_letter(WEEK_1ST_COL_NUM + 52 - 1)
WEEK_LST_COL_NUM = column_index_from_string(WEEK_LST_COL)


class Project:
    """Represents a Project to be justified"""

    def __init__(self, name: str, start: date, end: date):
        self.name = name
        self.name_coord = None
        self.start = start
        self.end = end
        self.respect_hints = False
        self.interval_limits = []
        self.projected = 0
        self.projected_coord = None
        self.justified = 0
        self.justified_eoy = 0

    def __str__(self) -> str:
        return self.__repr__()

    def __repr__(self) -> str:
        # , i={['{:.2f}'.format(i) for i in self.interval_limits]}
        return f"<PRJ {self.name} ({_f(self.start)}/{_f(self.end)}, eoi/eoy/projected={self.justified_eoi}/{self.justified_eoy}/{self.projected})>"

    def is_open(self, working_day: date) -> bool:
        """Returns whether a working day falls within a project

        Args:
            working_day (date): the day to test

        Returns:
            bool: True if the working day falls within a project
        """
        return self.start <= working_day and working_day <= self.end

    def is_full(self, current_interval: int) -> bool:
        """Returns whether the project has used up all its projected dedication

        Returns:
            bool: True if no more time can be dedicated to the project
        """
        # TODO: Not true when using week interval
        # assert 1 <= current_interval and current_interval <= 12, "current_interval must be a datetime month (from 1 to 12)"
        return ((self.justified_eoi >= self.projected) or (self.justified_eoi >= self.interval_limits[current_interval-1]))

    def get_availability(self, available_today: int, granularity: int, current_interval: int) -> int:
        """Returns the maximum number of minute that can be dedicated by the project taking into account different limits.

        Args:
            limit_daily (int): the maximum number of minutes that can be daily dedicated to all projects
            granularity (int): the minimum interval for justification
            current_interval (int): the current justification interval. Used to determine the limit for that interval

        Case: return project.available 
        +--------------> interval_limit
        +------------>   available_today
        +         +      granular_span
        +--->            project.available

        Case: return project.available 
        +--------------> interval_limit
        +----->          available_today
        +         +      granular_span
        +--->            project.available

        Case: return available_today
        +--------------> interval_limit
        +----->          available_today
        +         +      granular_span
        +----------->    project.available

        Case: return granular_span
        +--------------> interval_limit
        +------------>   available_today
        +         +      granular_span
        +----------->    project.available

        Case: return granular_span
        +--------------> interval_limit
        +------------>   available_today
        +         +      granular_span
        +----------->    project.available

        Case: return interval_limit
        +------->        interval_limit
        +------------>   available_today
        +         +      granular_span
        +----------->    project.available

        Returns:
            int: the number of minutes
        """
        # TODO: Not true when using week interval
        # assert 1 <= current_interval and current_interval <= 12, "current_month must be a datetime month (from 1 to 12)"
        if current_interval == 12 and self.name == 'TREASURE _ WP7':
            logger.debug("Should debug...")
        available_this_year = self.projected - self.justified_eoy
        available_this_month = self.interval_limits[current_interval -
                                                    1] - self.justified_eoi
        granular_span = max(math.ceil(min(
            [available_this_year, available_this_month, available_today]) / granularity), 1) * granularity
        return min([available_this_year, available_this_month, available_today, granular_span])


def _f(object: date) -> str:
    if type(object) == type(date.today()) or type(object) == type(datetime.now()):
        return object.strftime(DEFAULT_FMT)
    return str(object)


def _get_cell_coordinate(worksheet, cell_range: str) -> str:
    return f"{quote_sheetname(worksheet.title)}!{absolute_coordinate(cell_range)}"


def _load_project_justification_hints(worksheet: Worksheet, prj_row:int, justification_interval:str, justification_year:int, working_days: list) -> dict:
    from_month = justification_interval == MONTH_INTERVAL
    translate = False
    ws_min_col = MONTH_1ST_COL_NUM if from_month else WEEK_1ST_COL_NUM
    ws_max_col = MONTH_LST_COL_NUM if from_month else WEEK_LST_COL_NUM
    
    # Figure out whether it is a week or month report
    interval_1st_header:str = worksheet[f"{get_column_letter(ws_min_col)}{INTERVAL_HEADER_ROW}"].value
    if from_month:
        if interval_1st_header.lower() != date.today().replace(month=1).strftime("%B").lower():
            logger.warning(f"Worksheet '{worksheet.title}' is not a monthly report. Assuming a weekly report.")
            ws_min_col = WEEK_1ST_COL_NUM
            ws_max_col = WEEK_LST_COL_NUM
            translate = True
    else:
        if not interval_1st_header.lower().startswith("semana"):
            logger.warning(f"Worksheet '{worksheet.title}' is not a weekly report. Assuming a monthly report, data loss might occur.")
            ws_min_col = MONTH_1ST_COL_NUM
            ws_max_col = MONTH_LST_COL_NUM
            translate = True

    # Read the justified time for each interval
    prj_hints = {}
    for interval_col in worksheet.iter_cols(min_col=ws_min_col,
                                            max_col=ws_max_col,
                                            min_row=prj_row,
                                            max_row=prj_row):
        for interval_cell in interval_col:
            # NOTE: internally work with minutes...
            prj_hints[interval_cell.column - ws_min_col] = round(interval_cell.value * 60)

    
    if translate:
        translated_hints = { i:0 for i in range (0, 12 if from_month else 52) }
        # Take all working days determine each input interval and output interval
        # NOTE: We do not take into account the start and end date of the project
        #   as it has not been read yet... We however try to get the most accurate
        #   translation when going from month to week (using the number of working
        #   days in that month)
        days_in_interval = { i:0 for i in range (0, 12) }
        def _cumulate(working_day):
            days_in_interval[working_day.month-1] += 1
        list(map(_cumulate, working_days))
        
        for working_day in working_days:
            if from_month:
                intput_interval = working_day.isocalendar().week - 1
                output_interval = working_day.month - 1
                translated_hints[output_interval] += prj_hints[intput_interval]
            else:
                intput_interval = working_day.month - 1
                output_interval = working_day.isocalendar().week - 1
                translated_hints[output_interval] += prj_hints[intput_interval] / days_in_interval[intput_interval]
        prj_hints = { k: int(round(v)) for k, v in translated_hints.items() }
    return prj_hints


def _load_justification_hints(worksheet: Worksheet, only_projects: list, justification_interval: str, justification_year:int, working_days: list) -> dict:
    """Returns existing justification for project found in a report worksheet

    Args:
        worksheet (Worksheet): the worksheet containing existing justification
            for a list of project. Must have been generated by this script.
        only_projects (list): a list of project to filter on
        justification_interval (str): indicating whether the justification 
            interval is week or month.

    Returns:
        dict: a map { project_name -> { interval (zero_based) -> justified_eoi } }
    """
    hints = {}
    # Look up the project name in column PRJ_NAME_COL from PRJ_1ST_ROW
    for prj_col in worksheet.iter_cols(min_col=PRJ_NAME_COL_NUM, max_col=PRJ_NAME_COL_NUM, min_row=PRJ_1ST_ROW):
        for prj_cell in prj_col:
            project_name = prj_cell.value
            if not project_name:
                # Stop at the first empty row
                break
            if only_projects and project_name not in only_projects:
                # And filter out skipped projects
                logger.debug(
                    f"Skipping hints for project '{project_name}' ...")
                continue
            logger.debug(f"Loading hints for project '{project_name}' ...")
            hints[project_name] = _load_project_justification_hints(worksheet, prj_cell.row, justification_interval, justification_year, working_days)
    logger.debug(f"hints={pformat(hints)}")
    return hints


def _load_project_base_info(worksheet: Worksheet, employee: str, project_row: int, name_col_idx: int, start_col_idx: int, end_col_idx: int, year_columns: list, justification_year: int, only_projects: list):
    name_cell = worksheet.cell(row=project_row, column=name_col_idx)

    if only_projects and name_cell.value not in only_projects:
        logger.debug(f"Skipping config for project '{name_cell.value}' ...")
        return None

    start_cell = worksheet.cell(row=project_row, column=start_col_idx)
    end_cell = worksheet.cell(row=project_row, column=end_col_idx)
    projected_cell = worksheet.cell(
        row=project_row, column=year_columns[justification_year])
    if not projected_cell.value:
        projected_cell.value = 0.0
    assert type(start_cell.value) == type(datetime.now(
    )), f"Invalid type detected for cell '{employee}!{start_cell.coordinate}'. Expected a date."
    assert type(end_cell.value) == type(datetime.now(
    )), f"Invalid type detected for cell '{employee}!{end_cell.coordinate}'. Expected a date."
    assert type(projected_cell.value) == type(0.0) or type(projected_cell.value) == type(
        0), f"Invalid type detected for cell '{employee}!{projected_cell.coordinate}'. Expected a float or an int."
    project = Project(str(name_cell.value), start_cell.value, end_cell.value)
    project.name_coord = _get_cell_coordinate(worksheet, name_cell.coordinate)
    # NOTE: internally work with minutes...
    project.projected = projected_cell.value * 60
    project.projected_coord = _get_cell_coordinate(
        worksheet, projected_cell.coordinate)
    return project


def _get_project_working_days(project, working_days):
    project_working_days = list(
        filter(lambda d: project.start <= d and d <= project.end, working_days))
    if len(project_working_days) <= 0:
        raise ValueError(f"Invalid project breakdown: project '{project.name}' "
                         f"has projected dedication ({int(project.projected)}'), "
                         f"but no working days. It is configured to start on the "
                         f"{_f(project.start)} and finish on the {_f(project.end)}.")
    return project_working_days


def _get_project_daily_average(project, project_working_days, max_daily_limit):
    project_daily_average = project.projected / len(project_working_days)
    if project_daily_average > max_daily_limit:
        raise ValueError(f"Invalid project breakdown: project '{project.name}' "
                         f"has daily average ({project_daily_average}) which is "
                         f"greater than the max daily limit ({max_daily_limit})")
    return project_daily_average


def _get_interval_dates(justification_year: int, justification_interval: str, i: int):
    if justification_interval == MONTH_INTERVAL:
        assert 0 <= i and i < 12, "i must be between 0 and 12 when using {justification_interval} interval"
        start = datetime(justification_year, i+1, 1, 0, 0)
        end = start + relativedelta(months=+1, days=-1)
    elif justification_interval == WEEK_INTERVAL:
        assert 0 <= i and i < 52, "i must be between 0 and 52 when using {justification_interval} interval"
        first_weekday = datetime(justification_year, 1, 1, 0, 0)
        if first_weekday.weekday() > 4:
            first_weekday += relativedelta(days=7-first_weekday.weekday())
        start = first_weekday + relativedelta(days=7*i)
        end = first_weekday + relativedelta(days=7*i+6)
    else:
        raise ValueError(
            f"Invalid '--justification-interval' parameter. Must be either 'month' or 'week'.")
    return start, end


def _get_interval_limit(days, start, end, daily_average):
    interval_days = list(filter(lambda d: start <= d and d <= end, days))
    return len(interval_days) * daily_average


def _set_interval_limit(project, project_working_days, hints, start, end, current_interval, project_daily_average):
    interval_limit = int(round(_get_interval_limit(
        project_working_days, start, end, project_daily_average)))
    # BUG: The hints are on a month basis. this is not working when using week interval
    hint = hints.get(project.name, {}).get(current_interval, -1)
    if hint != -1:
        # If an hint was specified, *must* respect it.
        logger.debug(
            f"Overriding default dedication for project {project.name} for {_f(start)}: from {interval_limit} to {hint}")
        project.respect_hints = True
        project.interval_limits[current_interval] = hint
    else:
        project.interval_limits[current_interval] = interval_limit


def _load_project_breakdown(project_breakdown_filename: str,
                            employee: str,
                            only_projects: str,
                            justification_interval: str,
                            use_justification_hints: bool,
                            justification_year: int,
                            working_days: list,
                            defaults) -> dict:
    """Returns a dict containing information about the projects for the given employee

    Args:
        project_breakdown_filename (str): the filename of the Excel file containing the project breakdown
        employee (str): the name of the employee
        only_projects (str): a comma separated list of project name to filter on.
        justification_interval (str): indicates whether justification interval is 'month' or 'week'
        use_justification_hints (bool): whether to use the hints found in a possible already existing justification sheet
        justification_year (int): the year for which the justification is to be produced.
        working_days (list(date)): a sorted list of working days as per the established working calendar.
        defaults (dict): the default read from the 'default.yaml' file

    Raises:
        ValueError: raised if the employee does not have a project breakdown

    Returns:
        A dict containg the projects details:
        { project_name -> Project (project_name, start, end, interval_limits, justified, projected)

    """
    logger.info(
        f"Loading projects breakdown from {project_breakdown_filename} for '{employee}' ...")
    workbook = openpyxl.load_workbook(project_breakdown_filename)
    if employee not in workbook:
        raise ValueError(
            f"Invalid '--employee' parameter: No worksheet '{employee}' found in {project_breakdown_filename}.")

    only_projects = only_projects.split(",") if only_projects else None

    hints = {}
    if use_justification_hints:
        hint_worksheet_name = f"{employee} - {justification_year}"
        if hint_worksheet_name in workbook:
            logger.debug(
                f"Will use justification hints found in worksheet '{hint_worksheet_name}' ...")
            hints = _load_justification_hints(workbook[hint_worksheet_name], 
                                              only_projects, 
                                              justification_interval, 
                                              justification_year,
                                              working_days)
        else:
            logger.warning(
                f"You requested to use hints, but no worksheet named '{hint_worksheet_name}' was found. Skipping hints altogether.")

    worksheet = workbook[employee]
    # determining the column indexes for each relevant column
    name_col_idx = 0
    start_col_idx = 0
    end_col_idx = 0
    year_columns = {}
    logger.debug(f"Determining column indices ...")
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == defaults['project-name-column']:
                name_col_idx = cell.column
            elif cell.value == defaults['project-start-column']:
                start_col_idx = cell.column
            elif cell.value == defaults['project-end-column']:
                end_col_idx = cell.column
            elif (isinstance(cell.value, int) or isinstance(cell.value, str)) and int(cell.value) >= 2000 and int(cell.value) <= 2100:
                # The column seems to be a year...
                year_columns[int(cell.value)] = cell.column
            elif cell.value:
                logger.warning(
                    f"Ignoring unknown column with header {cell.column} (='{cell.value}' /{type(cell.value)}) in worksheet {employee}.")

    # NOTE: internally work with minutes...
    max_daily_limit = float(defaults['max-daily-limit'])*60

    logger.debug(f"Extracting projects config ...")
    # The map { project_name -> Project }
    projects = {}

    # Each row is a project ...
    for i in range(worksheet.min_row+1, worksheet.max_row+1):

        project = _load_project_base_info(worksheet, employee, i, name_col_idx,
                                          start_col_idx, end_col_idx, year_columns, justification_year, only_projects)
        if not project:
            # Reached the end of the list of project
            continue

        # The project has some projected dedication for the justification_year
        if project.projected > 0:
            logger.debug(f"Loading config for project '{project.name}' ...")
            project_working_days = _get_project_working_days(project, working_days)
            project_daily_average = _get_project_daily_average(project, project_working_days, max_daily_limit)

            if justification_interval == MONTH_INTERVAL:
                number_of_interval = 12
            elif justification_interval == WEEK_INTERVAL:
                number_of_interval = 52
            else:
                raise ValueError(
                    f"Invalid '--justification-interval' parameter. Must be either 'month' or 'week'.")

            # Calculate the limit for each interval
            project.interval_limits = [0] * number_of_interval
            # Start from the first weekday of the year...
            for i in range(0, number_of_interval):
                start, end = _get_interval_dates(
                    justification_year, justification_interval, i)
                if not project.is_open(start) and not project.is_open(end):
                    logger.debug(
                        f"Project '{project.name}' is not opened in interval {_f(start)} / {_f(end)}. Skipping hints ...")
                    continue
                _set_interval_limit(
                    project, project_working_days, hints, start, end, i, project_daily_average)
            projects[project.name] = project
        else:
            logger.debug(
                f"Skipping project '{project.name}': not opened in any interval of {justification_year}.")

    return projects


def _get_justifiation_dates(interval, period):
    """Return the justification start and end date

    Args:
        interval (str): a string indicating whether the justification should be a month or a week
        period (date): the start date of the justification period

    Returns:
        date: the start date for the justification period
        date: the end date for the justification period
    """
    if interval == MONTH_INTERVAL:
        start = period.replace(day=1)
        end = period.replace(day=calendar.monthrange(
            start.year, start.month)[1])
    else:
        start = period - timedelta(days=period.weekday())
        end = period + timedelta(days=6-period.weekday())
    return start, end


def _is_last_business_day_in_month(working_day: date) -> bool:
    last_day = working_day.replace(
        max(calendar.monthcalendar(working_day.year, working_day.month)[-1][:5]))
    return last_day == working_day


def _get_working_days(justification_start, holidays):
    """Returns the list of working day for the justification year removing the specified holidays

    Args:
        justification_start (datetime.date): the start of the justification period. Usually the 1st of the month to justify
        holidays (list(datetime.date)): a list of date containing the holidays

    Returns:
        list(date): the list of working days
    """
    # ok, I always justify a full year and then extract the intersting period.
    # This is due to the fact that we do not want to justify periods that are not a multiple of the
    # justification granulariry. Only during the last period of the year, is this allowed.
    logger.debug(
        f"Creating the holiday calendar for Aragon in {justification_start.year} ...")
    work_calendar = Aragon()
    # With the holiday calendar I create the list of the opened days in the justification year
    holidays = holidays.split(",")
    january_first = justification_start.replace(month=1, day=1)

    working_days = []
    for delta_day in (range(1, 367 if calendar.isleap(justification_start.year) else 366)):
        target_day = january_first + timedelta(days=delta_day)
        if work_calendar.is_working_day(target_day) and _f(target_day) not in holidays:
            working_days.append(target_day)
    return working_days, work_calendar


def _set_cell_to_title(cell: Cell, title: str, defaults:dict):
    cell.value = title
    try:
        cell.style = defaults['project-report-interval-title-style']
    except:
        logger.warning(
            f"Invalid configuration 'project-report-interval-title-style': Missing named style '{defaults['project-report-interval-title-style']}' in Excel file. Check your defaults.yaml.")
    

def _set_cell_to_time(cell: Cell, minutes: float):
    """Set the value of the given cell 

    Args:
        cell (Cell): _description_
        minutes (float): _description_
    """
    if PREFERRED_FMT == 'decimal':
        cell.value = minutes/60
        cell.number_format = "0.0"
    else:
        # Excel processes time entries as a decimal fraction of a day
        cell.value = minutes/(60 * 24)
        cell.number_format = "[h]:mm"


def _set_cell_to_project_name(cell: Cell, project: str, defaults: dict):
    if defaults['project-report-use-reference']:
        cell.value = f"={project.name_coord}"
    else:
        cell.value = project.name
    cell.protection = Protection(locked=True)
    try:
        cell.style = defaults['project-report-project-name-style']
    except:
        logger.warning(
            f"Invalid configuration 'project-report-project-name-style': Missing named style '{defaults['project-report-project-name-style']}' in Excel file. Check your defaults.yaml.")


def _set_cell_to_projected_dedication(cell: Cell, project: Project, defaults: dict):
    if defaults['project-report-use-reference']:
        cell.value = f"={project.projected_coord}"
    else:
        _set_cell_to_time(cell, project.projected)
    cell.number_format = CELL_FMT
    cell.protection = Protection(locked=True)


def _set_cell_to_justified_dedication(cell: Cell, named_range_name: str, defaults: dict):
    cell.value = f"=SUM({named_range_name})"
    cell.number_format = CELL_FMT


def _set_cell_to_total(cell: Cell, named_range_name: str, defaults: dict):
    cell.value = f"=SUM({named_range_name})"
    cell.number_format = CELL_FMT


def _safe_named_range_name(name: str) -> str:
    return re.sub('[^0-9a-zA-Z_]+', '', name).lower()


def _add_named_range(worksheet, name: str, cell_range: str) -> str:
    named_range = DefinedName(_safe_named_range_name(
        name), attr_text=_get_cell_coordinate(worksheet, cell_range))
    worksheet.defined_names.add(named_range)
    return named_range


def _get_interval_col(interval_1st_col: str, i: int) -> int:
    assert interval_1st_col == WEEK_1ST_COL or interval_1st_col == MONTH_1ST_COL
    return get_column_letter(column_index_from_string(interval_1st_col)+i-1)


def _pick_best_open_project(open_projects, current_interval, daily_available, granularity):
    """Returns the best project to fit into a time slot.

    The best project is defined as follow:
    1. Project that have respect_hints set to True take precedence.
    2. Project with the biggest available time slot come first
    3. project with soonest end date come first
    4. Alphabetical order on project name

    Args:
        open_projects (list): list of project
        current_interval (int): the current interval id
        daily_available (int): the time left available for the day
        granularity (int): the preferred time interval granularity

    Returns:
        tuple: a tuple containing the project and the available time
    """
    # respect_hints, available_time, end date, project name
    # NOTE: in order to reverse sort by end project date, we need to
    #   create a pivot date
    pivot = date(2000, 1, 1)
    by_availability = list(filter(lambda t: t[2] != 0, list(map(lambda project: (project, project.respect_hints, project.get_availability(
        daily_available, granularity, current_interval), pivot - project.end.date(), project.name), open_projects))))
    # logger.debug(f"by_availability={pformat(by_availability, compact=True)}")
    if len(by_availability) == 0:
        return None, None
    by_availability_sorted = sorted(
        by_availability, key=lambda x: (x[1], x[2], x[3], x[4]))
    # logger.debug(f"by_availability_sorted={pformat(by_availability_sorted, compact=True)}")
    t = by_availability_sorted[-1]
    return t[0], t[2]


def _get_interval(interval:str, d:datetime):
    if interval == MONTH_INTERVAL:
        return d.month
    if interval == WEEK_INTERVAL:
        return d.isocalendar().week
    assert False


def _is_new_interval(interval: str, d: datetime, i: int):
    month = d.month
    week = d.isocalendar().week
    if not i:
        return True, month if interval == MONTH_INTERVAL else week
    if (interval == MONTH_INTERVAL and i != month):
        return True, month
    if (interval == WEEK_INTERVAL and i != week):
        return True, week
    return False, i


def _generate_report(employee, project_breakdown: str, projects: dict, working_days, daily_dedications, justification_year: int, justification_interval:str, defaults: dict):
    logger.info("Generating report ...")
    workbook = openpyxl.load_workbook(project_breakdown, keep_vba=True)
    if defaults['worksheet-report-template'] not in workbook:
        raise ValueError(
            f"Invalid configuration: No worksheet '{defaults['worksheet-report-template']}' found in {project_breakdown}.")
    report_template = workbook[defaults['worksheet-report-template']]

    report_title = f"{employee} - {justification_year}"
    if report_title in workbook:
        logger.warning(f"Removing existing report '{report_title}' ...")
        del workbook[report_title]

    report = workbook.copy_worksheet(report_template)
    report.title = report_title

    # Should try to use defined name instead...
    # definition = report.defined_names["employee"]
    report['B1'].value = employee
    report['B2'].value = employee
    report['B3'].value = justification_year
    report['B4'].value = defaults['max-yearly-limit']

    # Let's define some ranges
    _add_named_range(report, "employee", 'B2')
    _add_named_range(report, "justification_year", 'B3')
    _add_named_range(report, "max_yearly_limit", 'B4')

    # I massage the daily_dedications as { project_name -> {month -> justified} }
    justifications = {}
    for working_day, project_dedications in daily_dedications.items():
        for project_name, justified in project_dedications.items():
            if project_name not in justifications.keys():
                justifications[project_name] = {}
            interval = _get_interval(justification_interval, working_day)
            if interval not in justifications[project_name]:
                justifications[project_name][interval] = 0
            justifications[project_name][interval] += justified

    logger.debug(f"justifications={pformat(justifications)}")
    # TODO: Way too much magic numbers! Indices are way too fragile! How to improve templating?

    interval_1st_col = MONTH_1ST_COL if justification_interval == MONTH_INTERVAL else WEEK_1ST_COL
    interval_lst_col = MONTH_LST_COL if justification_interval == MONTH_INTERVAL else WEEK_LST_COL
    max_range = 12 if justification_interval == MONTH_INTERVAL else 52

    #
    # Header of the justification period
    #
    first = date.today().replace(year=justification_year, month=1, day=1)
    for i in range(1, max_range+1):
        interval_column = _get_interval_col(interval_1st_col, i)
        title = first.replace(month=i).strftime("%B") if justification_interval == MONTH_INTERVAL else f"Semana {i}"
        _set_cell_to_title(report[f"{interval_column}{INTERVAL_HEADER_ROW}"], title.title(), defaults)
    try:
        for cell in report[f"{INTERVAL_HEADER_ROW}:{INTERVAL_HEADER_ROW}"]:
            cell.style = defaults['project-report-interval-title-style']
    except: None

    #
    # The report for the projects
    #
    for i, project_name in enumerate(sorted(justifications.keys())):

        project_row = PRJ_1ST_ROW + i
        project = projects[project_name]

        report.insert_rows(project_row)

        # let's set a couple of ranges
        _add_named_range(report, f"prj_total_projected_{project_name}", f"{PROJECTED_COL}{project_row}")
        _add_named_range(report, f"prj_total_justified_{project_name}", f"{JUSTIFIED_COL}{project_row}")
        _add_named_range(report, f"prj_{justification_interval}_justified_{project_name}", f"{interval_1st_col}{project_row}:{interval_lst_col}{project_row}")

        # Set the project name
        _set_cell_to_project_name(report[f"{PRJ_NAME_COL}{project_row}"], project, defaults)

        # The 'Horas propuestas' for that project
        _set_cell_to_projected_dedication(report[f"{PROJECTED_COL}{project_row}"], project, defaults)

        # The 'Horas justificadas' for that project (utilizamos la SUM de Excel como comprobacion)
        _set_cell_to_justified_dedication(report[f"{JUSTIFIED_COL}{project_row}"], _safe_named_range_name(f"prj_{justification_interval}_justified_{project_name}"), defaults)

        # The justified time for each interval
        for i in range(1, max_range+1):
            justified = 0
            if i in justifications[project_name]:
                justified = justifications[project_name][i]
            interval_column = _get_interval_col(interval_1st_col, i)
            _set_cell_to_time(report[f"{interval_column}{project_row}"], justified)

    #
    # 'Total proyectos' row
    #
    # 'Total proyectos' / 'Horas propuestas'
    totals_row = project_row+2
    _add_named_range(report, "all_prj_projected", f"{PROJECTED_COL}{PRJ_1ST_ROW}:{PROJECTED_COL}{project_row}")
    _add_named_range(report, "all_prj_projected_total", f"{PROJECTED_COL}{totals_row}")
    _set_cell_to_total(report[f"{PROJECTED_COL}{totals_row}"], "all_prj_projected", defaults)

    # 'Total proyectos' / 'Horas justificadas'
    _add_named_range(report, "all_prj_justified", f"{JUSTIFIED_COL}{PRJ_1ST_ROW}:{JUSTIFIED_COL}{project_row}")
    _add_named_range(report, "all_prj_justified_total", f"{JUSTIFIED_COL}{totals_row}")
    _set_cell_to_total(report[f"{JUSTIFIED_COL}{totals_row}"], "all_prj_justified", defaults)

    # 'Total proyectos' / Month
    for i in range(1, max_range+1):
        interval_column = _get_interval_col(interval_1st_col, i)
        named_range_name = _safe_named_range_name(f"all_prj_{justification_interval}_justified_{i}")
        _add_named_range(report, named_range_name,f"{interval_column}{PRJ_1ST_ROW}:{interval_column}{project_row}")
        _set_cell_to_total(report[f"{interval_column}{totals_row}"], named_range_name, defaults)

    #
    # 'Limites mensuales' row
    #
    # We write the maximum number of hours for this month MAX_DAILY_LIMIT*monthly_working_days
    MAX_DAILY_LIMIT = float(defaults['max-daily-limit'])*60
    limits_row = totals_row+1
    for i in range(1, max_range+1):
        interval_column = _get_interval_col(interval_1st_col, i)
        interval_working_days = list(filter(lambda d: _get_interval(justification_interval, d) == i, working_days))
        interval_limit = MAX_DAILY_LIMIT * len(interval_working_days)
        _set_cell_to_time(report[f"{interval_column}{limits_row}"], interval_limit)

    try:
        for row in report[f"A{totals_row}:A{limits_row}"]:
            for cell in row:    
                cell.style = defaults['project-report-interval-title-style']
    except: None
        

    alert_color = defaults['project-report-alert-color']
    warning_color = defaults['project-report-warning-color']
    ok_color = defaults['project-report-ok-color']
    grey_color = defaults['project-report-grey-color']

    # Doing some more formatting
    # Set the project's monthly dedication in grey if it is 0
    fill = PatternFill(bgColor=grey_color)
    rule = CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{interval_1st_col}{PRJ_1ST_ROW}:{interval_lst_col}{project_row}", rule)

    # BEWARE: The conditional formatting syntax is slightly different from the formula syntax. It uses ',' instead of ';'
    # Set the monthly justified dedication in red/yellow/green if it breaches the monthly limit (row offset = 1)
    fill = PatternFill(bgColor=alert_color)
    rule = FormulaRule(formula=['INDIRECT("RC",FALSE) > OFFSET(INDIRECT("RC",FALSE),1,0)'], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{interval_1st_col}{totals_row}:{interval_lst_col}{totals_row}", rule)

    fill = PatternFill(bgColor=warning_color)
    rule = FormulaRule(formula=['INDIRECT("RC",FALSE) = OFFSET(INDIRECT("RC",FALSE),1,0)'], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{interval_1st_col}{totals_row}:{interval_lst_col}{totals_row}", rule)

    fill = PatternFill(bgColor=ok_color)
    rule = FormulaRule(formula=['INDIRECT("RC",FALSE) < OFFSET(INDIRECT("RC",FALSE),1,0)'], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{interval_1st_col}{totals_row}:{interval_lst_col}{totals_row}", rule)

    # Set the project's total justified dedication in red/yellow/green if it breaches the projected limit (col offset = -1)
    fill = PatternFill(bgColor=alert_color)
    rule = FormulaRule(formula=['INDIRECT("RC",FALSE) > OFFSET(INDIRECT("RC",FALSE),0,-1)'], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{JUSTIFIED_COL}{PRJ_1ST_ROW}:{JUSTIFIED_COL}{project_row}", rule)

    fill = PatternFill(bgColor=warning_color)
    rule = FormulaRule(formula=['INDIRECT("RC",FALSE) = OFFSET(INDIRECT("RC",FALSE),0,-1)'], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{JUSTIFIED_COL}{PRJ_1ST_ROW}:{JUSTIFIED_COL}{project_row}", rule)

    fill = PatternFill(bgColor=ok_color)
    rule = FormulaRule(formula=['INDIRECT("RC",FALSE) < OFFSET(INDIRECT("RC",FALSE),0,-1)'], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{JUSTIFIED_COL}{PRJ_1ST_ROW}:{JUSTIFIED_COL}{project_row}", rule)

    # Set the total projected dedication in red if it breaches the yearly limit, green otherwise
    fill = PatternFill(bgColor=alert_color)
    rule = FormulaRule(formula=[f"${PROJECTED_COL}${totals_row} > max_yearly_limit"], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{PROJECTED_COL}{totals_row}", rule)
    fill = PatternFill(bgColor=ok_color)
    rule = FormulaRule(formula=[f"${PROJECTED_COL}${totals_row} <= max_yearly_limit"], stopIfTrue=True, fill=fill)
    report.conditional_formatting.add(f"{PROJECTED_COL}{totals_row}", rule)

    del workbook[defaults['worksheet-report-template']]

    try:
        workbook.save(defaults['project-report'])
    except PermissionError:
        logger.error(
            f"Can't write '{defaults['project-report']}'. Is it open in Excel?")


def distribute_hours(args):
    locale.setlocale(locale.LC_ALL, DEFAULT_LOCALE)

    # We interanlly work with minutes.
    MAX_DAILY_LIMIT = float(args.defaults['max-daily-limit'])*60
    MAX_YEARLY_LIMIT = float(args.defaults['max-yearly-limit'])*60
    granularity = float(args.justification_granularity)*60
    # Some sanity checks ...
    assert (MAX_DAILY_LIMIT >=
            granularity), f"Invalid configuration: daily limit ({int(MAX_DAILY_LIMIT)}') must be superior to justification granularity ({int(granularity)}')."
    assert (MAX_DAILY_LIMIT %
            granularity) == 0, f"Invalid configuration: daily limit ({int(MAX_DAILY_LIMIT)}') must be a multiple of the justification granularity ({int(granularity)}')."
    assert (MAX_YEARLY_LIMIT %
            granularity) == 0, f"Invalid configuration: yearly limit ({int(MAX_YEARLY_LIMIT)}') must be a multiple of the justification granularity ({int(granularity)}')."

    justification_start, justification_end = _get_justifiation_dates(
        args.justification_interval, args.justification_period)
    justification_year = justification_start.year
    holidays = ','.join(
        [args.defaults['university-holidays'], args.employee_holidays])
    working_days, work_calendar = _get_working_days(
        justification_start, holidays)

    # Load project breakdown for employee ...
    projects = _load_project_breakdown(args.project_breakdown,
                                       args.employee,
                                       args.only_projects,
                                       args.justification_interval,
                                       args.use_justification_hints,
                                       justification_year,
                                       working_days,
                                       args.defaults)

    logger.info(f"Justifying '{args.justification_interval}' interval "
                f"from '{_f(justification_start)}' to '{_f(justification_end)}' ...")
    logger.debug(
        f"{str(justification_year)} has {len(working_days)} working day(s)")

    # A map of working days with the project and the corresponding justified time
    daily_dedications = {working_day: {} for working_day in working_days}

    current_interval = None
    for working_day in working_days:
        # Reset the project justified dedication when changing interval (month/week)
        is_new_interval, new_interval = _is_new_interval(
            args.justification_interval, working_day, current_interval)
        if is_new_interval:
            current_interval = new_interval
            logger.debug(f"New {args.justification_interval} ({current_interval}) detected. "
                         "Resetting projects EOI dedication ...")
            list(map(lambda p: setattr(p, 'justified_eoi', 0), projects.values()))
            logger.debug(f"projects={pformat(list(projects.values()))}")

        # List projects that still need to be justified for that day.
        # NOTE: that a project can be opened in the middle of the year.
        open_projects = list(filter(lambda p: p.is_open(
            working_day) and not p.is_full(current_interval), projects.values()))
        if len(open_projects) == 0:
            logger.debug(f"No project left opened on {_f(working_day)}")
            continue
        # daily_dedication: time already justified for that day
        daily_dedication = 0
        while daily_dedication < MAX_DAILY_LIMIT:
            # Find the project that has the biggest available time slot for the day.
            # Projects that have respect_hints set to True take precedence.
            project, available = _pick_best_open_project(
                open_projects, current_interval, MAX_DAILY_LIMIT - daily_dedication, granularity)
            if not project:
                logger.debug(f"No more time can be justified on {_f(working_day)}."
                             "Skipping to next day.")
                break
            logger.debug(
                f"Adding {available}' to project {project.name} on {_f(working_day)}")

            daily_dedication += available
            project.justified_eoi += available
            project.justified_eoy += available
            daily_dedications[working_day][project.name] = available

            assert project.justified_eoy <= project.projected, f"Project justified dedication to date ({int(project.justified_eoy)}') is superior to projected dedication ({int(project.projected)}') for project {project.name}"

    # I check whether some projects were left with some projected time in each interval.
    # That would be a signal that not all hours could be squeezed in the given interval.
    open_projects = list(filter(lambda p: p.is_open(
        working_day) and not p.is_full(current_interval), projects.values()))
    if len(open_projects) != 0 and _is_last_business_day_in_month(working_day):
        logger.warning(
            f"Reached the end of the month with some project left opened {_f(working_day)}")

    logger.debug(f"daily_dedications={pformat(daily_dedications)}")

    _generate_report(args.employee,
                     args.project_breakdown,
                     projects,
                     working_days,
                     daily_dedications,
                     justification_year,
                     args.justification_interval,
                     args.defaults)


def parse_command_line(defaults):
    parser = ArgumentParser(prog='distribute_hours',
                            description=__doc__, formatter_class=RawTextHelpFormatter)
    parser.add_argument(
        '--debug', action="store_true", help='Run the program in debug', required=False, default=False)
    parser.add_argument(
        '--quiet', action="store_true", help='Run the program in quiet mode', required=False, default=False)
    parser.add_argument(
        '--project-breakdown', help=f"The Excel file containing the project breakdown. By default '{defaults['project-breakdown']}'.", required=False, default=defaults['project-breakdown'])
    parser.add_argument(
        '--employee', help=f"The employee to justify hours for. Note that the project breakdown file must contain a worksheet whose name is equal to the employee's name. By default '{defaults['employee']}'.", required=False, default=defaults['employee'])
    parser.add_argument(
        '--justification-period', type=lambda s: datetime.strptime(s, DEFAULT_FMT), help=f"The justification period in '{DEFAULT_FMT}' format. By default today's date.", required=False, default=_f(date.today()))
    parser.add_argument(
        '--justification-interval', help=f"The justification interval. Either 'month' or 'week'. By default '{defaults['justification-interval']}'.", required=False, default=defaults['justification-interval'])
    parser.add_argument(
        '--justification-granularity', type=float, help=f"The justification granularity in hour. By default '{defaults['justification-granularity']}'.", required=False, default=defaults['justification-granularity'])
    parser.add_argument(
        '--employee-holidays', help=f"A comma separated list of extra holidays to take into account (using '{DEFAULT_FMT}' format). First week of January by default.", required=False, default=defaults['employee-holidays'])
    parser.add_argument(
        '--only-projects', help=f"A comma separated list of projects to take into account. All project if not specified.", required=False, default=None)
    parser.add_argument(
        '--use-justification-hints', action="store_true", help='Use a previously created justification worksheet to calculate projected dedication. The justified dedication found in said worksheet will take precedence on the projected dedication. The hint worksheet must have been created by this program in order to be readable.', required=False, default=False)
    parser.set_defaults(func=distribute_hours)
    return parser.parse_args()


def main():
    with open(os.path.join(os.path.dirname(__file__), 'defaults.yaml'), 'r', encoding='utf-8') as f:
        defaults = yaml.safe_load(f)
    args = parse_command_line(defaults)
    args.defaults = defaults
    try:
        if args.debug:
            coloredlogs.install(level=logging.DEBUG, logger=logger)
        elif args.quiet:
            coloredlogs.install(level=logging.ERROR, logger=logger)
        else:
            coloredlogs.install(level=logging.INFO, logger=logger)
        return args.func(args)
    except Exception as e:
        logging.error(e, exc_info=True)
        return 1


if __name__ == '__main__':
    sys.exit(main())
